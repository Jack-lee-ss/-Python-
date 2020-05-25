# -*- coding: utf-8 -*-
import re
import time

import openpyxl
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium import webdriver
from lxml import etree
a=r'C:\Program Files (x86)\Google\Chrome\Application\chromedriver.exe'
driver=webdriver.Chrome(executable_path=a)

def parse_url(url,x,s):
	j=1
	infos={}
	# Location=(By.XPATH,'/html/body/div[5]/div/div[1]/div/h4')  ## 注：这里xpath 只能定位元素或者标签，不能定位到文本内容。
	# WebDriverWait(driver,timeout=5,ignored_exceptions=None).until(EC.presence_of_element_located(Location))
	html=etree.HTML(driver.page_source)
	time.sleep(3)
	job_name=html.xpath('/html/body/div[5]/div/div[1]/div/h4/text()')[0]
	name=html.xpath('/html/body/div[5]/div/div[1]/div/h1/text()')[0]
	job_request_salary=html.xpath('/html/body/div[5]/div/div[1]/dd/h3/span[1]/text()')[0]
	job_request_place=html.xpath('/html/body/div[5]/div/div[1]/dd/h3/span[2]/text()')[0]
	job_request_place=re.sub(r'[\s/]','',job_request_place).strip()
	job_request_expresion=html.xpath('/html/body/div[5]/div/div[1]/dd/h3/span[3]/text()')[0]
	job_request_expresion=re.sub(r'[/]','',job_request_expresion).strip()
	job_request_education=html.xpath('/html/body/div[5]/div/div[1]/dd/h3/span[4]/text()')[0]
	job_request_education=re.sub(r'[/]','',job_request_education).strip()
	job_advvatage=html.xpath('//*[@id="job_detail"]/dd[1]/p/text()')[0]
	job_detail=html.xpath('//*[@id="job_detail"]/dd[2]/div/p/text()')
	job_details=''.join(job_detail).strip().replace(r'\xa0','')
	infos={
		'公司':job_name,
		'工作':name,
		'薪水':job_request_salary,
		'工作地点':job_request_place,
		'工作经验':job_request_expresion,
		'教育要求':job_request_education,
		'优点':job_advvatage,
		'要求':job_details
	}
	#print(infos)
	print('=========第 %s 页第 %s 个职位详情============='%(x,s))
	driver.close()
	time.sleep(1)
	save_details(infos)
	
	
def swipe_down():
	driver.maximize_window()
	h = driver.get_window_size()['height']
	for x in range(1, 5):
		js = "var q=document.documentElement.scrollTop=" + str((h/2) * x)
		driver.execute_script(js)
		time.sleep(1)


def save_details(details):
	book=openpyxl.Workbook()
	sheet=book.active
	sheet.title='拉钩网职位数据一览表'
	sheet.cell(1,1).value='职位信息'
	sheet.cell(1,3).value='数据'
	row=2
	#for x in details:
	for k,v in details.items():
		sheet.cell(row,1).value=k
		sheet.cell(row,3).value=v
		row+=1
	row+=1
	print('===== 已保存一份数据 =====')
	book.save('拉勾数据.xlsx')
	
	
def main():
	base_url = 'https://www.lagou.com/'
	driver.get(base_url)
	source = driver.page_source
	html = etree.HTML(source)
	driver.maximize_window()
	time.sleep(1)
	box=driver.find_element_by_id('cboxClose').click()
	time.sleep(1)
	driver.find_element_by_id('search_input').send_keys('python')
	driver.find_element_by_id('search_button').click()
	try:
		location = (By.CLASS_NAME, 'body-btn')
		WebDriverWait(driver, 10).until(EC.presence_of_element_located(location))
		driver.find_element_by_class_name('body-btn').click()
	except:
		pass
	time.sleep(1)
	
	x=1
	while True:
		s=1
		html = etree.HTML(driver.page_source)
		#urls=html.xpath('//ul[@class="item_con_list"]//a[@class="position_link"]//@href')
		urls=html.xpath('//a[@class="position_link"]/@href')
		for url in urls:
			driver.execute_script('window.open("%s")' % url)
			driver.switch_to.window(driver.window_handles[1])
			swipe_down()
			details=parse_url(url,x,s)
			driver.switch_to.window(driver.window_handles[0])
			s+=1
		x+=1
		next_btn = driver.find_element_by_xpath('//div[@class="pager_container"]/span[last()]')
		if 'pager_next pager_next_disabled' in next_btn.get_attribute('class'):
			break
		else:
			next_btn.click()
		time.sleep(2)
		
if __name__ == '__main__':
	main()
	
#  print(etree.tostring(urls, encoding='utf-8').decode('utf-8')) <Element ul at 0x20520cdd6c8>
# i=1
# 	# for url in url_list:
# 	# 	print(etree.tostring(url, encoding='utf-8').decode('utf-8'))
# 	# 	print('====== %s ==========='%i)
# 	# 	i+=1
# 	#print(url_list)