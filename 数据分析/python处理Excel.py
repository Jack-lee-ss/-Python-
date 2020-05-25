### ===================  处理Excel表格 =====================
from calendar import monthrange
from datetime import datetime, timedelta
import os
import time

import openpyxl
import xlrd
## xlrd库中 open_workbook打开Excel文件，并返回一个Book对象，这个对象代表打开的excel文件
## 获取 Excel 文件中表单(sheet) 的数量 和 所有 表单(sheet) 的名字
book = xlrd.open_workbook("D:/data/income.xlsx")
print(f"包含表单数量 {book.nsheets}")
print(f"表单的名分别为: {book.sheet_names()}")

# 表单索引从0开始，获取第一个表单对象
a=book.sheet_by_index(0)
print(a)

# 获取名为2018的表单对象
a=book.sheet_by_name('2018')
print(a)

# 获取所有的表单对象，放入一个列表返回
a=book.sheets()
print(a,'\n')

## 获取了表单对象后，可以根据其属性得到：
sheet = book.sheet_by_index(0)
print(f"表单名：{sheet.name} ")
print(f"表单索引：{sheet.number}")
print(f"表单行数：{sheet.nrows}")
print(f"表单列数：{sheet.ncols}")
print(sheet.name,'\n')

## 获取了表单对象后，可以使用 cell_value 方法，参数为行号和列号，读取指定单元格中的文本内容
# 行号、列号都是从0开始计算
print(f"单元格A1内容是: {sheet.cell_value(rowx=7, colx=0)}")
# 转化成整型数据
print(f"单元格A1内容是: {int(sheet.cell_value(rowx=4, colx=1))}")

# 还可以使用 row_values 方法，参数为行号，读取指定行所有单元格的内容，存放在一个列表中返回。
print(f'第3行的内容是：{sheet.row_values(rowx=3)}')
print(f'第3行的内容是：{sheet.row_values(rowx=4)}')
# 还可以使用 col_values 方法，参数为列号，读取指定列所有单元格的内容，存放在一个列表中返回.
print(f'第2列中第4行到第6行的内容是：{sheet.col_values(colx=1,start_rowx=3,end_rowx=6)}','\n')

### 汇总数据（去除带*号不规范数据，然后求和）
# 获取一个工作簿中多少个表，并且赋给变量,输出内存地址
sheets=book.sheets()
incomes_3years=0
for sheet in sheets:
	
	# 每一个表中第二列第二行的数据为收入
	incomes=sheet.col_values(colx=1,start_rowx=1)
	
	# 去掉包含星号的月份收入
	toSubstract = 0
	
	# 月份累计
	months=sheet.col_values(colx=0)
	
	# 提取不规范的月份及收入，求和
	for row,month in enumerate(months):
		if type(month) is str and month.endswith('*'):
			income=sheet.cell_value(row,1)
			print(row,income)
			toSubstract+=income
	
	realincomes=int(sum(incomes)-toSubstract)
	print(f'{sheet.name}真实总收入：{realincomes}')
	print(f'{sheet.name}不规范收入：{toSubstract}')
	print('------------------------','\n')
incomes_3years+=realincomes
print(f'总收入：{incomes_3years}')
print(realincomes)


##########  Excel数据的写入与修改
## 创建一个Excel工作簿
book=openpyxl.Workbook()
# 创建时会自动产生一个sheet,这个表是自动生成的，通过active激活
sh=book.active
# # 修改当前的标题为‘电影详情表’
sh.title='电影详情表'
## 创建不同工作表sheets，所有表没有写出序号的，默认排在最后。序号从0开始
sheet1=book.create_sheet('水电费详情表',0)
sheet2=book.create_sheet('生活详情表',1)
sheet3=book.create_sheet('日常支出详情表',2)

# 根据名称获取某个sheet对象，确定某一个表
sh = book['生活详情表']

# 在该表上给任一个单元格写入内容
sh['C8'] = '你好'

# 获取该表上某个单元格内容,如果没有数据则返回 None
print(sh['A1'].value)

# 根据行号,列号， 给第一个单元格写入内容，和表中序列号对应。
# 注意和 xlrd 不同，是从 1 开始
book['日常支出详情表'].cell(2,2).value = '哈哈哈'

### 输入一个字典内容到表中
s={'译名': '决X中途岛/中途岛战役/中途岛海战',
   '片名': '中间的way',
   '年代': '2019',
   '产地': '中国/美国',
   '类别': '战争/历史',
   '语言': '英语',
   '字幕': '中文字幕',
   'IMDb评分': '6.9/10 from 12377 用户',
   '视频尺寸': '1920 x 798',
   '片长': '138 Mins',
   '导演': '罗兰·艾默里奇 Roland Emmerich',
   '主演': '伍迪·哈里森 Woody Harrelson 帕特里克·威尔森 Patrick Wilson 卢克·伊万斯 Luke Evans 艾德·斯克林 Ed Skrein 丹尼斯·奎德 DennisQuaid 曼迪·摩尔 Mandy Moore 亚历山大·路德韦格 Alexander Ludwig 艾伦·艾克哈特 Aaron Eckhart 达伦·克里斯 Darren Criss'}
print(len(s))

book['电影详情表'].cell(1,1).value='详情'
book['电影详情表'].cell(1,2).value='数据'
# 确定从第2行开始写入
row=2
for i,j in s.items():
	book['电影详情表'].cell(row, 1).value = i
	book['电影详情表'].cell(row, 2).value = j
	row+=1

row=1
x=0
for i, j in s.items():
	book['电影详情表'].cell(row, x+1).value = i
	book['电影详情表'].cell(row+1, x+1).value = j
	x+=1
	
# 保存文件,一般在最后一步。运行代码需要关闭工作簿
# book.save('信息.xlsx')


##### 文件操作
## 创建目录  os.makedirs 可以递归的创建目录结构
# 运行代码即可创建；exist_ok=True 指定了，如果某个要创建的目录已经存在，也不报错
os.makedirs('D:/data/new/temp/cn',exist_ok=True)

####  时间和日期：获取当前时间对应的数字；使用time库中time模块
## time.time() 会返回 从 1970年1月1日0点到当前时间的经过的秒数，可以简称为秒数时间。
before = time.time()
s={'译名': '决X中途岛/中途岛战役/中途岛海战',
   '片名': '中间的way',
   '年代': '2019',
   '产地': '中国/美国',
   '类别': '战争/历史',
   '语言': '英语',
   '字幕': '中文字幕',
   'IMDb评分': '6.9/10 from 12377 用户',
   '视频尺寸': '1920 x 798',
   '片长': '138 Mins',
   '导演': '罗兰·艾默里奇 Roland Emmerich',
   '主演': '伍迪·哈里森 Woody Harrelson 帕特里克·威尔森 Patrick Wilson 卢克·伊万斯 Luke Evans 艾德·斯克林 Ed Skrein 丹尼斯·奎德 DennisQuaid 曼迪·摩尔 Mandy Moore 亚历山大·路德韦格 Alexander Ludwig 艾伦·艾克哈特 Aaron Eckhart 达伦·克里斯 Darren Criss'}
row = 1
x = 0
for i, j in s.items():
	book['电影详情表'].cell(row, x + 1).value = i
	book['电影详情表'].cell(row + 1, x + 1).value = j
	x += 1
#保存文件,一般在最后一步。运行代码需要关闭工作簿
book.save('信息.xlsx')

after = time.time()
print(f"输入数据的花费时间{after-before}")
print(time.strftime('%Y%m%d %H:%M:%S',time.localtime(after-before)),'\n')


### 指定格式字符串显示时间
## 要得到 当前时间 对应的字符串，可以这样实现：
# 默认格式：2020-02-08 13:09:15.418233
print(str(datetime.now()))

# 如果要指定输出的时间格式，可以像下面这样,连接符自选
print(datetime.now().strftime('%Y~%m~%d ==== %H:%M:%S'))

# 当然，也可以使用time库来格式化显示字符串
time.strftime('%Y-%m-%d %H:%M:%S',time.localtime())

## 数字表示的时间转化为字符串表示
#  如果要将某个指定秒数时间（从epoch时间点开始计算），而不是当前时间，转化为字符串格式
print(time.strftime('%Y%m%d %H:%M:%S',time.localtime(time.time())) )

## 字符串时间转化为整数时间:1970 年到 2015-08-01 23:59:59 多少秒
print(int(time.mktime(time.strptime('2015-08-01 23:59:59', '%Y-%m-%d %H:%M:%S'))))

## 获取某个时间 对应 的年月日时分秒数字，可以使用datatime库
print(datetime.now())
print(datetime.now().month)
print(datetime.now().day)
print(datetime.now().year)
print(datetime.now().second)
print(datetime.now().hour)
print(datetime.now().minute,'\n')

### 推算时间 ：获得指定时间字符串对应星期几；使用 datetime类的 strptime方法(变秒)
theday=datetime.strptime('2020-02-08','%Y-%m-%d')
print(theday.weekday())
print(f'星期 {theday.weekday()}')
#### 不准确


### 从某个时间点往前或者后推 一段时间
a=datetime.strptime('2020-02-08','%Y-%m-%d')
b=a+timedelta(days=100)
print(b)
print(b.weekday())
b=a-timedelta(days=20)
print(b)

## 获取某一个月共有多少天 calendar 模块里面的 函数
mr=monthrange(2020,1)
print(mr[1])


