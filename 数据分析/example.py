import openpyxl
from lxml import etree
import time
import requests

a=time.time()
headers = {
	'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36'
}


# def get_urls(url):
# 	reponse = requests.get(url, headers=headers)
# 	## 查看源码后，该网站用 gbk 编码，所以 reponse.content.decode 自己解码在编码
# 	text = reponse.content.decode(encoding='gbk', errors='ignore')
# 	html = etree.HTML(text)
# 	## 提取页面的 url 地址
# 	urls = html.xpath('//div[@class="co_content8"]//b/a/@href')
# 	urls = map(lambda url: 'https://www.dytt8.net' + url, urls)
# 	return urls


def parse_detail_url(url):
	response = requests.get(url=url, headers=headers)
	text = response.content.decode(encoding='gbk', errors='ignore')
	# print(text)
	html = etree.HTML(text)
	## 提取详情页信息
	infos = html.xpath('//div[@id="Zoom"]')[0]
	texts = infos.xpath('.//text()')
	movie = {}
	for index, text in enumerate(texts):
		
		if text.startswith('◎译　　名'):
			movie['译名'] = text.replace('◎译　　名', '').strip()
		
		if text.startswith('◎片　　名'):
			movie['片名'] = text.replace('◎片　　名', '').strip()
		
		elif text.startswith('◎年　　代'):
			movie['年代'] = text.replace('◎年　　代', '').strip()
		
		elif text.startswith('◎产　　地'):
			movie['产地'] = text.replace('◎产　　地', '').strip()
		
		elif text.startswith('◎类    别'):
			movie['类别'] = text.replace('◎类    别', '').strip()
		
		elif text.startswith('◎语　　言'):
			movie['语言'] = text.replace('◎语　　言', '').strip()
		
		elif text.startswith('◎字　　幕'):
			movie['字幕'] = text.replace('◎字　　幕', '').strip()
		
		elif text.startswith('◎上映日期'):
			movie['上映日期'] = text.replace('◎上映日期', '').strip()
		
		elif text.startswith('◎豆瓣评分'):
			movie['豆瓣评分'] = text.replace('◎豆瓣评分', '').strip()
		
		elif text.startswith('◎导　　演'):
			movie['导演'] = text.replace('◎导　　演', '').strip()
		
		elif text.startswith('◎编　　剧'):
			a = text.replace('◎编　　剧', '').strip()
			a_s = [a]
			for x in range(index + 1, len(texts)):
				b = texts[x].strip()
				if b.startswith('◎主　　演'):
					break
				a_s.append(b)
			list1 = [str(i) for i in a_s]
			list2 = ''.join(list1)
			movie['编剧'] = list2
		
		elif text.startswith('◎主　　演'):
			a = text.replace('◎主　　演', '').strip()
			a_s = [a]
			for x in range(index + 1, len(texts)):
				b = texts[x].strip()
				if b.startswith('◎标　　签'):
					break
				a_s.append(b)
			list1 = [str(i) for i in a_s]
			list2 = ''.join(list1)
			movie['主演'] = list2
		
		elif text.startswith('◎简　　介'):
			a_s = []
			for x in range(index + 1, len(texts)):
				b = texts[x].strip()
				if b.startswith('◎'):
					break
				elif b.startswith('【下载地址】'):
					break
				a_s.append(b)
			list1 = [str(i) for i in a_s]
			list2 = ''.join(list1)
			movie['简介'] = list2
	
		# elif text.startswith('◎获奖情况 '):
		# 	a_s=[]
		# 	for x in range(index+1,len(texts)):
		# 		b_list = texts[x].strip()
		# 		if text.startswith('【下载地址】'):
		# 			break
		# 		a_s.append(b_list)
		# 	list1 = [str(i) for i in a_s]
		# 	list2 = ''.join(list1)
		# 	movie['获奖情况'] = list2
		
	downloads = html.xpath('//div[@class="co_area2"]//a/@href')[0]
	movie['下载地址'] = downloads
	return movie


def write_detail(s):
	book = openpyxl.Workbook()
	sh = book.active
	sh.title = '电影排列方式一'
	sh1 = book.create_sheet('电影排列方式二')
	
	row = 2
	book['电影排列方式一'].cell(1, 1).value = '详情列'
	book['电影排列方式一'].cell(1, 2).value = '数据'
	for i,j in s.items():
		book['电影排列方式一'].cell(row, 1).value = i
		book['电影排列方式一'].cell(row, 2).value = j
		row+=1
	row+=1
	
	book.save('电影天堂信息.xlsx')


def spider():
	base_url = 'https://www.dytt8.net/html/gndy/dyzz/20200202/59664.html'
	mov = parse_detail_url(base_url)
	result= write_detail(mov)
	b=time.time()
	print(b - a)
if __name__ == '__main__':
	spider()
	